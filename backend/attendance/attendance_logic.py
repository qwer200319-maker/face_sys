"""
attendance_logic.py
===================
Core business logic for:
  1. Auto-detect Check-In / Check-Out by time window
  2. Ignore scans during work hours + break time
  3. Calculate work hours automatically
  4. Daily summary per employee (In/Out/Work Hours/Late/OT)
"""
from datetime import datetime, timedelta, time as dtime
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# SECTION 1 — TIME WINDOW HELPERS
# ═══════════════════════════════════════════════════════════════

def _time_to_minutes(t: dtime) -> int:
    """Convert time to total minutes since midnight."""
    return t.hour * 60 + t.minute


def _minutes_to_time(mins: int) -> dtime:
    """Convert minutes (can be negative or >1440) back to time."""
    mins = mins % 1440  # wrap around midnight
    return dtime(mins // 60, mins % 60)


def _add_minutes(t: dtime, delta: int) -> dtime:
    return _minutes_to_time(_time_to_minutes(t) + delta)


def _in_window(t: dtime, start: dtime, end: dtime) -> bool:
    """
    Check if time t is within [start, end].
    Handles overnight wrap (e.g. 23:00 → 01:00).
    """
    ts, te, tt = _time_to_minutes(start), _time_to_minutes(end), _time_to_minutes(t)
    if ts <= te:
        return ts <= tt <= te        # normal window
    else:
        return tt >= ts or tt <= te  # overnight wrap


def get_scan_status(shift, now: datetime) -> str:
    """
    Determine what this scan means based on current time vs shift windows.

    Returns:
        'check_in'  → record as check in
        'check_out' → record as check out
        'break'     → inside break time, ignore
        'ignore'    → inside work hours or outside all windows, ignore
    """
    if not shift:
        # No shift → always check_in (simple mode)
        return 'check_in'

    t = now.time()

    # ── Priority 1: Break time → ignore ───────────────────────
    if shift.break_start and shift.break_end:
        if _in_window(t, shift.break_start, shift.break_end):
            return 'break'

    # ── Priority 2: Check-In window ───────────────────────────
    # Window: (start_time - checkin_before) → (start_time + checkin_after)
    # Example: 08:00 shift, before=30, after=120
    #          → 07:30 to 10:00 = check_in window
    ci_start = _add_minutes(shift.start_time, -shift.checkin_before)
    ci_end   = _add_minutes(shift.start_time,  shift.checkin_after)
    if _in_window(t, ci_start, ci_end):
        return 'check_in'

    # ── Priority 3: Check-Out window ──────────────────────────
    # Window: (end_time - checkout_before) → (end_time + checkout_after)
    # Example: 17:00 shift, before=30, after=120
    #          → 16:30 to 19:00 = check_out window
    co_start = _add_minutes(shift.end_time, -shift.checkout_before)
    co_end   = _add_minutes(shift.end_time,  shift.checkout_after)
    if _in_window(t, co_start, co_end):
        return 'check_out'

    # ── Priority 4: Everything else → ignore ──────────────────
    # This covers: mid-work hours, outside any window
    return 'ignore'


def is_late(shift, checkin_dt: datetime) -> tuple:
    """
    Check if check-in is late.
    Returns: (is_late: bool, late_minutes: int)
    """
    if not shift:
        return False, 0
    deadline = datetime.combine(checkin_dt.date(), shift.start_time)
    deadline += timedelta(minutes=shift.grace_minutes)
    if checkin_dt > deadline:
        mins = int((checkin_dt - deadline).total_seconds() / 60)
        return True, mins
    return False, 0


def is_in_cooldown(cooldowns: dict, emp_id: str, status: str, now: datetime, cooldown_sec: int) -> bool:
    """
    Check if this employee+status is within cooldown period.
    Prevents duplicate check_in or duplicate check_out records.
    """
    key  = f"{emp_id}_{status}"
    last = cooldowns.get(key)
    if last is None:
        return False
    return (now - last).total_seconds() < cooldown_sec


def update_cooldown(cooldowns: dict, emp_id: str, status: str, now: datetime):
    cooldowns[f"{emp_id}_{status}"] = now


# ═══════════════════════════════════════════════════════════════
# SECTION 2 — WORK HOURS CALCULATION
# ═══════════════════════════════════════════════════════════════

def calculate_work_hours(checkin_dt: datetime, checkout_dt: datetime, shift=None) -> dict:
    """
    Calculate work hours between check-in and check-out.
    Deducts break time automatically if shift has break configured.

    Returns dict:
        gross_hours  : total time from in to out
        break_hours  : break duration (if configured)
        net_hours    : gross - break (actual work)
        ot_hours     : net_hours - shift.duration_hours (if positive)
        is_undertime : net_hours < shift.duration_hours
    """
    if not checkin_dt or not checkout_dt:
        return {'gross_hours': 0, 'break_hours': 0, 'net_hours': 0, 'ot_hours': 0, 'is_undertime': False}

    total_seconds = (checkout_dt - checkin_dt).total_seconds()
    if total_seconds <= 0:
        return {'gross_hours': 0, 'break_hours': 0, 'net_hours': 0, 'ot_hours': 0, 'is_undertime': False}

    gross_hours = round(total_seconds / 3600, 2)
    break_hours = 0.0

    # Deduct break time if shift has break configured
    if shift and shift.break_start and shift.break_end:
        bs = _time_to_minutes(shift.break_start)
        be = _time_to_minutes(shift.break_end)
        if be > bs:
            break_hours = round((be - bs) / 60, 2)

    net_hours = round(gross_hours - break_hours, 2)
    net_hours = max(net_hours, 0)

    # Overtime = net work beyond expected shift duration
    ot_hours    = 0.0
    is_undertime = False
    if shift:
        expected = float(shift.duration_hours)
        if net_hours > expected:
            ot_hours = round(net_hours - expected, 2)
        elif net_hours < expected:
            is_undertime = True

    return {
        'gross_hours':  gross_hours,
        'break_hours':  break_hours,
        'net_hours':    net_hours,
        'ot_hours':     ot_hours,
        'is_undertime': is_undertime,
    }


def save_checkout_with_hours(employee_obj, checkout_dt: datetime, camera=None, confidence=0.0):
    """
    Save check_out record and:
      - Calculate work hours
      - Auto-create/update OvertimeRecord if OT exists
    Returns the saved AttendanceRecord.
    """
    from attendance.models import AttendanceRecord, OvertimeRecord

    # Find today's check-in
    checkin_rec = AttendanceRecord.objects.filter(
        employee=employee_obj,
        status='check_in',
        timestamp__date=checkout_dt.date(),
        is_unknown=False,
    ).order_by('-timestamp').first()

    checkin_dt = checkin_rec.timestamp if checkin_rec else None
    hours_data = calculate_work_hours(checkin_dt, checkout_dt, employee_obj.shift)

    # Save check_out record
    rec = AttendanceRecord.objects.create(
        employee   = employee_obj,
        camera     = camera,
        status     = 'check_out',
        confidence = confidence,
        is_unknown = False,
        work_hours = Decimal(str(hours_data['net_hours'])),
    )

    # Also update check_in record with work_hours
    if checkin_rec:
        checkin_rec.work_hours = Decimal(str(hours_data['net_hours']))
        checkin_rec.save(update_fields=['work_hours'])

    # Auto-create Overtime if OT exists
    if hours_data['ot_hours'] > 0 and checkin_dt:
        OvertimeRecord.objects.update_or_create(
            employee = employee_obj,
            date     = checkout_dt.date(),
            defaults = {
                'ot_start': checkin_dt,
                'ot_end':   checkout_dt,
                'ot_hours': Decimal(str(hours_data['ot_hours'])),
                'status':   'pending',
            }
        )
        logger.info(
            "OT created for %s: %.2fh",
            employee_obj.employee_id,
            hours_data['ot_hours']
        )

    return rec, hours_data


# ═══════════════════════════════════════════════════════════════
# SECTION 3 — DAILY SUMMARY PER EMPLOYEE
# ═══════════════════════════════════════════════════════════════

def get_employee_daily_summary(employee_obj, target_date) -> dict:
    """
    Build a complete daily summary for one employee on target_date.

    Returns:
        {
          employee_id, name, department, shift_name,
          check_in_time, check_out_time,
          is_late, late_minutes,
          gross_hours, break_hours, net_hours, ot_hours,
          is_undertime, status
        }
    """
    from attendance.models import AttendanceRecord

    records = AttendanceRecord.objects.filter(
        employee   = employee_obj,
        timestamp__date = target_date,
        is_unknown = False,
    ).order_by('timestamp')

    checkin_rec  = records.filter(status='check_in').first()
    checkout_rec = records.filter(status='check_out').last()

    checkin_dt  = checkin_rec.timestamp  if checkin_rec  else None
    checkout_dt = checkout_rec.timestamp if checkout_rec else None

    hours_data = calculate_work_hours(checkin_dt, checkout_dt, employee_obj.shift)

    late_flag, late_min = (False, 0)
    if checkin_dt and employee_obj.shift:
        late_flag, late_min = is_late(employee_obj.shift, checkin_dt)

    # Determine overall status
    if not checkin_rec:
        day_status = 'absent'
    elif not checkout_rec:
        day_status = 'present_no_out'  # checked in but not out yet
    elif hours_data['is_undertime']:
        day_status = 'undertime'
    elif hours_data['ot_hours'] > 0:
        day_status = 'overtime'
    else:
        day_status = 'present'

    return {
        'employee_id':   employee_obj.employee_id,
        'name':          employee_obj.name,
        'name_kh':       employee_obj.name_kh,
        'department':    employee_obj.department.name if employee_obj.department else '',
        'shift_name':    str(employee_obj.shift) if employee_obj.shift else '',
        'check_in_time': checkin_dt.strftime('%H:%M:%S')  if checkin_dt  else '—',
        'check_out_time':checkout_dt.strftime('%H:%M:%S') if checkout_dt else '—',
        'is_late':        late_flag,
        'late_minutes':   late_min,
        'gross_hours':    hours_data['gross_hours'],
        'break_hours':    hours_data['break_hours'],
        'net_hours':      hours_data['net_hours'],
        'ot_hours':       hours_data['ot_hours'],
        'is_undertime':   hours_data['is_undertime'],
        'status':         day_status,
    }


def get_daily_report(target_date, department_id=None) -> list:
    """
    Generate full attendance report for ALL employees on target_date.
    Optionally filter by department.
    Returns list of daily summary dicts.
    """
    from attendance.models import Employee
    employees = Employee.objects.filter(is_active=True).select_related('department', 'shift')
    if department_id:
        employees = employees.filter(department_id=department_id)

    results = []
    for emp in employees:
        summary = get_employee_daily_summary(emp, target_date)
        results.append(summary)

    # Sort: absent last, late first, then alphabetical
    order = {'absent': 3, 'present_no_out': 2, 'undertime': 1, 'overtime': 0, 'present': 0}
    results.sort(key=lambda x: (order.get(x['status'], 3), x['name']))
    return results


# ═══════════════════════════════════════════════════════════════
# SECTION 4 — EXCEL / PDF EXPORT HELPERS
# ═══════════════════════════════════════════════════════════════

def export_daily_report_excel(target_date, department_id=None):
    """
    Generate Excel workbook bytes for daily attendance report.
    Includes: In/Out times, work hours, late info, OT, status.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = get_daily_report(target_date, department_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance {target_date}"

    # ── Styles ────────────────────────────────────────────────
    hdr_fill   = PatternFill('solid', fgColor='1F4E79')
    hdr_font   = Font(bold=True, color='FFFFFF', size=11)
    alt_fill   = PatternFill('solid', fgColor='EBF3FB')
    late_fill  = PatternFill('solid', fgColor='FFF3CD')
    abs_fill   = PatternFill('solid', fgColor='FCE8E8')
    ot_fill    = PatternFill('solid', fgColor='E8F8F5')
    center     = Alignment(horizontal='center', vertical='center')
    thin       = Border(
        left  =Side(style='thin', color='D0D0D0'),
        right =Side(style='thin', color='D0D0D0'),
        top   =Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    # ── Title row ─────────────────────────────────────────────
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value     = f"Attendance Report — {target_date}"
    title_cell.font      = Font(bold=True, size=14, color='1F4E79')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # ── Summary row ───────────────────────────────────────────
    present = sum(1 for d in data if d['status'] != 'absent')
    absent  = sum(1 for d in data if d['status'] == 'absent')
    late    = sum(1 for d in data if d['is_late'])
    ot_cnt  = sum(1 for d in data if d['ot_hours'] > 0)

    ws.merge_cells('A2:M2')
    ws['A2'].value = (
        f"Total: {len(data)}   ✅ Present: {present}   "
        f"❌ Absent: {absent}   ⏰ Late: {late}   ⏱ OT: {ot_cnt}"
    )
    ws['A2'].font      = Font(bold=True, size=11, color='334155')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 22

    # ── Headers ───────────────────────────────────────────────
    headers = [
        '#', 'Emp ID', 'Name', 'Khmer Name', 'Department', 'Shift',
        'Check-In', 'Check-Out', 'Gross(h)', 'Break(h)', 'Net(h)',
        'OT(h)', 'Late', 'Late(min)', 'Status'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = thin
    ws.row_dimensions[3].height = 18

    # ── Data rows ─────────────────────────────────────────────
    STATUS_LABELS = {
        'present':         '✅ Present',
        'present_no_out':  '🟡 No Check-Out',
        'absent':          '❌ Absent',
        'undertime':       '⚠️ Undertime',
        'overtime':        '⏱️ OT',
    }

    for row_i, d in enumerate(data, 4):
        row_data = [
            row_i - 3,
            d['employee_id'], d['name'], d['name_kh'],
            d['department'], d['shift_name'],
            d['check_in_time'], d['check_out_time'],
            d['gross_hours'], d['break_hours'], d['net_hours'],
            d['ot_hours'],
            'Yes' if d['is_late'] else 'No',
            d['late_minutes'],
            STATUS_LABELS.get(d['status'], d['status']),
        ]

        # Row background by status
        if d['status'] == 'absent':
            row_fill = abs_fill
        elif d['is_late']:
            row_fill = late_fill
        elif d['ot_hours'] > 0:
            row_fill = ot_fill
        elif row_i % 2 == 0:
            row_fill = alt_fill
        else:
            row_fill = None

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row_i, col, val)
            cell.alignment = center
            cell.border    = thin
            if row_fill:
                cell.fill = row_fill
            # Highlight late column red
            if col == 13 and val == 'Yes':
                cell.font = Font(color='CC0000', bold=True)
            # Highlight OT column green
            if col == 12 and isinstance(val, float) and val > 0:
                cell.font = Font(color='006600', bold=True)

    # ── Column widths ─────────────────────────────────────────
    col_widths = [4, 10, 18, 18, 14, 22, 10, 10, 9, 9, 9, 8, 6, 10, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_daily_report_pdf(target_date, department_id=None):
    """
    Generate PDF bytes for daily attendance report.
    Landscape A4 with color-coded rows.
    """
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    data    = get_daily_report(target_date, department_id)
    buf     = io.BytesIO()
    doc     = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        leftMargin=1*cm,  rightMargin=1*cm,
    )
    styles  = getSampleStyleSheet()
    story   = []

    # Title
    story.append(Paragraph(
        f"<b>Attendance Report — {target_date}</b>",
        ParagraphStyle('title', fontSize=16, textColor=colors.HexColor('#1F4E79'),
                       spaceAfter=4, alignment=1)
    ))

    # Summary
    present = sum(1 for d in data if d['status'] != 'absent')
    absent  = sum(1 for d in data if d['status'] == 'absent')
    late    = sum(1 for d in data if d['is_late'])
    story.append(Paragraph(
        f"Total: {len(data)}  |  ✅ Present: {present}  |  "
        f"❌ Absent: {absent}  |  ⏰ Late: {late}",
        ParagraphStyle('sub', fontSize=10, textColor=colors.HexColor('#64748b'),
                       spaceAfter=10, alignment=1)
    ))

    # Table data
    headers = [
        '#', 'Emp ID', 'Name', 'Department', 'Shift',
        'Check-In', 'Check-Out', 'Net(h)', 'OT(h)', 'Late', 'Status'
    ]
    table_data = [headers]
    row_colors = []  # (row_idx, color)

    STATUS_COLORS = {
        'absent':          colors.HexColor('#FCE8E8'),
        'present_no_out':  colors.HexColor('#FFF9E6'),
        'undertime':       colors.HexColor('#FFF3CD'),
        'overtime':        colors.HexColor('#E8F8F5'),
        'present':         colors.white,
    }
    STATUS_LABELS = {
        'present':        'Present',
        'present_no_out': 'No Out',
        'absent':         'Absent',
        'undertime':      'Undertime',
        'overtime':       'OT',
    }

    for i, d in enumerate(data, 1):
        row = [
            str(i),
            d['employee_id'], d['name'],
            d['department'], d['shift_name'],
            d['check_in_time'], d['check_out_time'],
            str(d['net_hours']),
            str(d['ot_hours']) if d['ot_hours'] > 0 else '—',
            f"+{d['late_minutes']}m" if d['is_late'] else 'OK',
            STATUS_LABELS.get(d['status'], d['status']),
        ]
        table_data.append(row)
        row_colors.append((i, STATUS_COLORS.get(d['status'], colors.white)))

    tbl = Table(table_data, repeatRows=1,
                colWidths=[1*cm, 2.2*cm, 3.8*cm, 3*cm, 4*cm,
                           2.2*cm, 2.2*cm, 1.8*cm, 1.8*cm, 1.6*cm, 2.2*cm])

    style_cmds = [
        # Header
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 9),
        # Data
        ('FONTSIZE',   (0,1), (-1,-1), 8),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#D0D0D0')),
        ('ROWHEIGHT',  (0,0), (-1,-1), 18),
    ]

    # Apply row background colors
    for row_idx, color in row_colors:
        style_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), color))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return buf
