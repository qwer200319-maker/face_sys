import base64, io, cv2, numpy as np, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, datetime, timedelta
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from .models import (Employee, AttendanceRecord, Camera, Department,
                     Shift, OvertimeRecord, LeaveRequest, FaceEmbedding)
from .serializers import (EmployeeSerializer, AttendanceRecordSerializer,
                           CameraSerializer, DepartmentSerializer, ShiftSerializer,
                           OvertimeRecordSerializer, LeaveRequestSerializer)
from .face_engine import face_engine


def _decode_img(b64: str):
    if ',' in b64:
        b64 = b64.split(',')[1]
    arr = np.frombuffer(base64.b64decode(b64), np.uint8)
    return cv2.imdecode(arr, cv2.IMREAD_COLOR)


# ── Departments ──────────────────────────────────────────────
class DepartmentViewSet(viewsets.ModelViewSet):
    queryset         = Department.objects.all()
    serializer_class = DepartmentSerializer


# ── Shifts ───────────────────────────────────────────────────
class ShiftViewSet(viewsets.ModelViewSet):
    queryset         = Shift.objects.all()
    serializer_class = ShiftSerializer


# ── Employees ────────────────────────────────────────────────
class EmployeeViewSet(viewsets.ModelViewSet):
    queryset         = Employee.objects.select_related('department', 'shift').all()
    serializer_class = EmployeeSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['department', 'shift', 'is_active']

    @action(detail=True, methods=['post'], url_path='register-face')
    def register_face(self, request, pk=None):
        emp = self.get_object()
        b64 = request.data.get('image')
        if not b64:
            return Response({'error': 'No image'}, status=400)
        img = _decode_img(b64)
        if img is None:
            return Response({'error': 'Invalid image'}, status=400)
        face_engine.initialize()
        ok, msg = face_engine.register_face(emp.employee_id, img)
        if ok:
            from django.conf import settings
            return Response({
                'message':    msg,
                'face_count': FaceEmbedding.objects.filter(employee=emp).count(),
                'max_angles': getattr(settings, 'MAX_EMBEDDINGS_PER_PERSON', 5),
            })
        return Response({'error': msg}, status=400)

    @action(detail=True, methods=['delete'], url_path='clear-face')
    def clear_face(self, request, pk=None):
        emp = self.get_object()
        FaceEmbedding.objects.filter(employee=emp).delete()
        emp.face_embedding = None
        emp.save(update_fields=['face_embedding'])
        face_engine.reload()
        return Response({'message': 'Face cleared'})

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        total = Employee.objects.filter(is_active=True).count()
        reg   = Employee.objects.filter(is_active=True, embeddings__isnull=False).distinct().count()
        return Response({'total': total, 'face_registered': reg, 'not_registered': total - reg})

    @action(detail=False, methods=['get'], url_path='mobile-list')
    def mobile_list(self, request):
        """Lightweight endpoint for mobile view"""
        emp_id = request.query_params.get('employee_id')
        if not emp_id:
            return Response({'error': 'employee_id required'}, status=400)
        try:
            emp = Employee.objects.select_related('department', 'shift').get(employee_id=emp_id)
        except Employee.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)

        today   = date.today()
        records = AttendanceRecord.objects.filter(employee=emp, timestamp__date=today).order_by('-timestamp')
        ot_this_month = OvertimeRecord.objects.filter(
            employee=emp,
            date__month=today.month,
            date__year=today.year,
            status='approved',
        )
        total_ot = sum(float(r.ot_hours) for r in ot_this_month)

        return Response({
            'employee':      EmployeeSerializer(emp, context={'request': request}).data,
            'today_records': AttendanceRecordSerializer(records, many=True, context={'request': request}).data,
            'monthly_ot_hours': round(total_ot, 2),
        })


# ── Attendance ───────────────────────────────────────────────
class AttendanceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset         = AttendanceRecord.objects.select_related(
                           'employee', 'employee__department', 'employee__shift', 'camera').all()
    serializer_class = AttendanceRecordSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['is_unknown', 'status', 'is_late']

    @action(detail=False, methods=['get'], url_path='today-summary')
    def today_summary(self, request):
        today   = date.today()
        known   = AttendanceRecord.objects.filter(timestamp__date=today, is_unknown=False)
        present = known.values('employee').distinct().count()
        total   = Employee.objects.filter(is_active=True).count()
        late    = known.filter(is_late=True).values('employee').distinct().count()
        return Response({
            'date':               today.isoformat(),
            'total_employees':    total,
            'present':            present,
            'absent':             total - present,
            'late':               late,
            'on_time':            present - late,
            'unknown_detections': AttendanceRecord.objects.filter(timestamp__date=today, is_unknown=True).count(),
            'attendance_rate':    round(present / total * 100, 1) if total else 0,
        })

    @action(detail=False, methods=['get'], url_path='live-feed')
    def live_feed(self, request):
        qs = self.get_queryset().order_by('-timestamp')[:60]
        return Response(AttendanceRecordSerializer(qs, many=True, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='shift-summary')
    def shift_summary(self, request):
        today = date.today()
        shifts = Shift.objects.all()
        result = []
        for s in shifts:
            emps    = Employee.objects.filter(shift=s, is_active=True).count()
            present = AttendanceRecord.objects.filter(
                timestamp__date=today, is_unknown=False, employee__shift=s
            ).values('employee').distinct().count()
            result.append({
                'shift_id':   s.id,
                'shift_name': s.name,
                'start_time': s.start_time.strftime('%H:%M'),
                'end_time':   s.end_time.strftime('%H:%M'),
                'color':      s.color,
                'total':      emps,
                'present':    present,
                'absent':     emps - present,
            })
        return Response(result)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        df = request.query_params.get('date_from', date.today().isoformat())
        dt = request.query_params.get('date_to',   date.today().isoformat())
        qs = AttendanceRecord.objects.filter(
            timestamp__date__gte=df, timestamp__date__lte=dt,
            is_unknown=False, employee__isnull=False,
        ).select_related('employee','employee__department','employee__shift','camera').order_by('timestamp')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        hf = PatternFill('solid', fgColor='1F4E79')
        hb = Font(bold=True, color='FFFFFF', size=11)
        af = PatternFill('solid', fgColor='EBF3FB')
        headers = ['#','Emp ID','Name','Name (KH)','Department','Shift','Date','Time','Camera','Conf%','Late','Late(min)']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h); cell.fill = hf; cell.font = hb
            cell.alignment = Alignment(horizontal='center')
        for i, r in enumerate(qs, 2):
            row = [i-1, r.employee.employee_id, r.employee.name, r.employee.name_kh,
                   r.employee.department.name if r.employee.department else '',
                   r.employee.shift.name if r.employee.shift else '',
                   r.timestamp.strftime('%Y-%m-%d'), r.timestamp.strftime('%H:%M:%S'),
                   r.camera.name if r.camera else '', round(r.confidence*100,1),
                   'Yes' if r.is_late else 'No', r.late_minutes]
            for c, v in enumerate(row, 1):
                cell = ws.cell(i, c, v)
                if i % 2 == 0: cell.fill = af
                if v == 'Yes': cell.font = Font(color='FF0000', bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        res = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename="attendance_{df}_{dt}.xlsx"'
        return res

    @action(detail=False, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        df = request.query_params.get('date_from', date.today().isoformat())
        dt = request.query_params.get('date_to',   date.today().isoformat())
        qs = AttendanceRecord.objects.filter(
            timestamp__date__gte=df, timestamp__date__lte=dt,
            is_unknown=False, employee__isnull=False,
        ).select_related('employee','employee__department','employee__shift','camera').order_by('timestamp')

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=cm, bottomMargin=cm)
        sty = getSampleStyleSheet()
        data = [['#','Emp ID','Name','Dept','Shift','Date','Time','Late','Conf%']]
        for i, r in enumerate(qs, 1):
            data.append([str(i), r.employee.employee_id, r.employee.name,
                         r.employee.department.name if r.employee.department else '-',
                         r.employee.shift.name if r.employee.shift else '-',
                         r.timestamp.strftime('%Y-%m-%d'), r.timestamp.strftime('%H:%M:%S'),
                         f"+{r.late_minutes}m" if r.is_late else 'On time',
                         f'{r.confidence*100:.1f}%'])
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR',(0,0),(-1,0),  colors.white),
            ('FONTNAME',(0,0),(-1,0),   'Helvetica-Bold'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#EBF3FB')]),
            ('GRID',(0,0),(-1,-1), 0.5, colors.grey),
            ('FONTSIZE',(0,0),(-1,-1),  9),
            ('ALIGN',(0,0),(-1,-1),     'CENTER'),
        ]))
        doc.build([Paragraph(f'Attendance Report  {df} → {dt}', sty['Title']), Spacer(1,.3*cm), tbl])
        buf.seek(0)
        res = HttpResponse(buf.read(), content_type='application/pdf')
        res['Content-Disposition'] = f'attachment; filename="attendance_{df}_{dt}.pdf"'
        return res


# ── Overtime ─────────────────────────────────────────────────
class OvertimeViewSet(viewsets.ModelViewSet):
    queryset         = OvertimeRecord.objects.select_related('employee','employee__department').all()
    serializer_class = OvertimeRecordSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['status', 'employee', 'date']

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        ot = self.get_object()
        ot.status      = 'approved'
        ot.approved_by = request.data.get('approved_by', 'Admin')
        if ot.ot_end:
            delta = (ot.ot_end - ot.ot_start).seconds / 3600
            ot.ot_hours = round(delta, 2)
        ot.save()
        from .notify import notify_overtime_approved
        notify_overtime_approved(ot.employee.name, float(ot.ot_hours), str(ot.date))
        return Response(OvertimeRecordSerializer(ot).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        ot = self.get_object()
        ot.status = 'rejected'
        ot.save()
        return Response(OvertimeRecordSerializer(ot).data)

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        month = int(request.query_params.get('month', date.today().month))
        year  = int(request.query_params.get('year',  date.today().year))
        qs    = OvertimeRecord.objects.filter(date__month=month, date__year=year, status='approved')
        total_hours = sum(float(r.ot_hours) for r in qs)
        total_pay   = sum(r.ot_pay for r in qs)
        return Response({
            'month': month, 'year': year,
            'total_records': qs.count(),
            'total_ot_hours': round(total_hours, 2),
            'total_ot_pay':   round(total_pay, 2),
        })

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_ot_excel(self, request):
        month = int(request.query_params.get('month', date.today().month))
        year  = int(request.query_params.get('year',  date.today().year))
        qs    = OvertimeRecord.objects.filter(date__month=month, date__year=year
                    ).select_related('employee','employee__department').order_by('date')
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Overtime'
        hf = PatternFill('solid', fgColor='1F4E79')
        hb = Font(bold=True, color='FFFFFF')
        headers = ['#','Emp ID','Name','Department','Date','OT Hours','Rate','OT Pay','Status']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h); cell.fill = hf; cell.font = hb
        for i, r in enumerate(qs, 2):
            ws.append([i-1, r.employee.employee_id, r.employee.name,
                       r.employee.department.name if r.employee.department else '',
                       str(r.date), float(r.ot_hours), float(r.ot_rate),
                       round(r.ot_pay, 2), r.status])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        res = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename="overtime_{year}_{month:02d}.xlsx"'
        return res


# ── Leave ────────────────────────────────────────────────────
class LeaveViewSet(viewsets.ModelViewSet):
    queryset         = LeaveRequest.objects.select_related('employee').all()
    serializer_class = LeaveRequestSerializer
    filter_backends  = [DjangoFilterBackend]
    filterset_fields = ['status', 'leave_type', 'employee']

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        leave = self.get_object()
        leave.status      = 'approved'
        leave.approved_by = request.data.get('approved_by', 'Admin')
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        leave = self.get_object()
        leave.status = 'rejected'
        leave.save()
        return Response(LeaveRequestSerializer(leave).data)


# ── Cameras ──────────────────────────────────────────────────
class CameraViewSet(viewsets.ModelViewSet):
    queryset         = Camera.objects.all()
    serializer_class = CameraSerializer


# ═══════════════════════════════════════════════════════════════
# Daily Report endpoint — uses attendance_logic
# ═══════════════════════════════════════════════════════════════
from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .attendance_logic import (
    get_daily_report,
    export_daily_report_excel,
    export_daily_report_pdf,
)


@api_view(['GET'])
def daily_report_json(request):
    """
    GET /api/daily-report/?date=2025-01-15&department=1
    Returns full In/Out/WorkHours/Late/OT per employee for given date.
    """
    from datetime import date
    date_str  = request.query_params.get('date', date.today().isoformat())
    dept_id   = request.query_params.get('department')
    try:
        target = date.fromisoformat(date_str)
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD'}, status=400)

    data = get_daily_report(target, department_id=dept_id)
    return Response({
        'date':    date_str,
        'count':   len(data),
        'present': sum(1 for d in data if d['status'] != 'absent'),
        'absent':  sum(1 for d in data if d['status'] == 'absent'),
        'late':    sum(1 for d in data if d['is_late']),
        'records': data,
    })


@api_view(['GET'])
def daily_report_excel(request):
    """GET /api/daily-report/excel/?date=2025-01-15"""
    from datetime import date
    date_str = request.query_params.get('date', date.today().isoformat())
    dept_id  = request.query_params.get('department')
    try:
        target = date.fromisoformat(date_str)
    except ValueError:
        return HttpResponse('Invalid date', status=400)

    buf = export_daily_report_excel(target, department_id=dept_id)
    res = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    res['Content-Disposition'] = f'attachment; filename="daily_report_{date_str}.xlsx"'
    return res


@api_view(['GET'])
def daily_report_pdf(request):
    """GET /api/daily-report/pdf/?date=2025-01-15"""
    from datetime import date
    date_str = request.query_params.get('date', date.today().isoformat())
    dept_id  = request.query_params.get('department')
    try:
        target = date.fromisoformat(date_str)
    except ValueError:
        return HttpResponse('Invalid date', status=400)

    buf = export_daily_report_pdf(target, department_id=dept_id)
    res = HttpResponse(buf.read(), content_type='application/pdf')
    res['Content-Disposition'] = f'attachment; filename="daily_report_{date_str}.pdf"'
    return res
